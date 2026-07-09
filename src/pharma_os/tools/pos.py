"""Local source-workbook probability-of-success lookup."""

from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from pharma_os.schemas import AssetIdentityOutput, ClinicalTrialRecord, MissingDataFlag, PoSOutput, SourceMetadata
from pharma_os.tools._due_diligence_common import DEFAULT_POS_WORKBOOK, json_scalar, missing, norm, slug, to_float


def lookup_pos(
    trial: ClinicalTrialRecord,
    asset: AssetIdentityOutput,
    *,
    workbook_path: str | None = None,
) -> tuple[PoSOutput, SourceMetadata]:
    """Lookup source-only PoS from the local workbook."""

    path = Path(workbook_path or os.getenv("PHARMA_OS_POS_WORKBOOK_PATH") or DEFAULT_POS_WORKBOOK)
    source = SourceMetadata(
        source_id=f"pos_workbook:{slug(path.name)}",
        title="Source-based probability of success workbook",
        url=None,
        provenance="Local Source_Based_PoS_Workbook.xlsx AllBenchmarks lookup",
        source_type="pos_workbook",
        version=path.name,
    )
    flags: list[MissingDataFlag] = []
    if not path.exists():
        return (
            PoSOutput(workbook_path=str(path), source_ids=(), missing_data_flags=(missing("pos-workbook-missing", "pos", "workbook_path", f"Workbook not found: {path}", "high"),), confidence=0.0),
            source,
        )
    phase = _phase_for_workbook(trial.phases)
    disease_area = _disease_area_for_workbook(asset.therapeutic_area, trial.conditions)
    if not phase:
        flags.append(missing("pos-phase-missing", "pos", "current_phase", "Trial phase could not be mapped to workbook labels.", "high"))
    if not disease_area:
        flags.append(missing("pos-disease-area-missing", "pos", "disease_area", "Disease area could not be mapped to workbook labels.", "high"))
    value = None
    row_out: dict[str, str | int | float | bool | None] = {}
    lookup_key = f"Disease Area|{disease_area}|{phase}" if disease_area and phase else None
    if lookup_key:
        try:
            wb = load_workbook(path, data_only=False, read_only=False)
            for row_map in _all_benchmarks_rows(wb):
                if norm(_row_lookup_key(row_map)) == norm(lookup_key):
                    value = _row_pos_value(row_map)
                    row_out = {str(k): json_scalar(v) for k, v in row_map.items() if k is not None}
                    break
            if value is None:
                flags.append(missing("pos-row-missing", "pos", "probability_of_success", f"No workbook row found for {lookup_key}.", "high"))
        except Exception as exc:
            flags.append(missing("pos-workbook-error", "pos", "probability_of_success", f"Workbook lookup failed: {exc.__class__.__name__}.", "high"))
    return (
        PoSOutput(
            probability_of_success=value,
            current_phase=phase,
            disease_area=disease_area,
            workbook_path=str(path),
            lookup_key=lookup_key,
            benchmark_row=row_out,
            source_ids=(source.source_id,) if value is not None else (),
            missing_data_flags=tuple(flags),
            confidence=0.9 if value is not None else 0.0,
        ),
        source,
    )


def _disease_area_for_workbook(therapeutic_area: str | None, conditions: tuple[str, ...]) -> str | None:
    text = " ".join([therapeutic_area or "", *conditions]).casefold()
    normalized = _normalize_words(text)
    if any(term in normalized for term in ("oncology", "cancer", "tumor", "glioblastoma", "malignancy", "carcinoma", "lymphoma", "leukemia")):
        return "Oncology"
    if any(term in normalized for term in ("neurology", "neurologic", "neurodegenerative", "alzheimer", "parkinson", "multiple sclerosis", "dementia", "epilepsy")):
        return "Neurology"
    if any(term in normalized for term in ("autoimmune", "immunology", "rheumatology", "lupus", "sle", "systemic lupus erythematosus", "rheumatoid arthritis", "psoriasis", "atopic dermatitis")):
        return "Autoimmune"
    if any(term in normalized for term in ("cardiology", "cardiovascular", "heart failure", "pulmonary hypertension", "hypertension")):
        return "Cardiovascular"
    if any(term in normalized for term in ("endocrine", "endocrinology", "diabetes", "thyroid")):
        return "Endocrine"
    if any(term in normalized for term in ("metabolic", "obesity", "dyslipidemia", "hyperlipidemia")):
        return "Metabolic"
    if any(term in normalized for term in ("respiratory", "pulmonary", "asthma", "copd")):
        return "Respiratory"
    if any(term in normalized for term in ("infectious", "infection", "viral", "bacterial", "hiv", "covid")):
        return "Infectious disease"
    if any(term in normalized for term in ("hematology", "hematologic", "anemia", "hemophilia", "thrombocytopenia")):
        return "Hematology"
    if any(term in normalized for term in ("gastroenterology", "gastrointestinal", "crohn", "ulcerative colitis", "ibd")):
        return "Gastroenterology"
    if any(term in normalized for term in ("ophthalmology", "ophthalmic", "retina", "macular", "glaucoma")):
        return "Ophthalmology"
    if any(term in normalized for term in ("psychiatry", "psychiatric", "depression", "schizophrenia", "bipolar")):
        return "Psychiatry"
    if any(term in normalized for term in ("urology", "urologic", "bladder", "prostate")):
        return "Urology"
    if any(term in normalized for term in ("allergy", "allergic", "urticaria")):
        return "Allergy"
    return None


def _phase_for_workbook(phases: tuple[str, ...]) -> str | None:
    text = " ".join(phases).upper()
    if "PHASE3" in text or "PHASE 3" in text:
        return "Phase III"
    if "PHASE2" in text or "PHASE 2" in text:
        return "Phase II"
    if "PHASE1" in text or "PHASE 1" in text or "EARLY_PHASE1" in text:
        return "Phase I"
    return None


def _all_benchmarks_rows(wb: Any) -> list[dict[str, Any]]:
    for ws in wb.worksheets:
        for table in ws.tables.values():
            if table.displayName in {"AllBenchmarks", "AllBenchmarksTable"}:
                return _table_rows(ws, table.ref)
    if "AllBenchmarks" in wb.sheetnames:
        return _sheet_rows(wb["AllBenchmarks"])
    raise ValueError("Workbook missing AllBenchmarks sheet/table")


def _table_rows(ws: Any, ref: str) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    headers = [str(ws.cell(row=min_row, column=col).value).strip() for col in range(min_col, max_col + 1)]
    rows: list[dict[str, Any]] = []
    for row_num in range(min_row + 1, max_row + 1):
        row = {headers[index]: ws.cell(row=row_num, column=min_col + index).value for index in range(len(headers))}
        if any(value is not None for value in row.values()):
            rows.append(row)
    return rows


def _sheet_rows(ws: Any) -> list[dict[str, Any]]:
    headers = [str(ws.cell(row=1, column=col).value).strip() for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_num in range(2, ws.max_row + 1):
        row = {headers[index]: ws.cell(row=row_num, column=index + 1).value for index in range(len(headers))}
        if any(value is not None for value in row.values()):
            rows.append(row)
    return rows


def _row_lookup_key(row: dict[str, Any]) -> str:
    direct = _get_value(row, "Lookup Key") or _get_value(row, "Key")
    if direct:
        return str(direct)
    return "|".join(
        [
            str(_get_value(row, "Benchmark Type") or _get_value(row, "Benchmark Group") or ""),
            str(_get_value(row, "Selected Benchmark") or _get_value(row, "Benchmark Name") or ""),
            str(_get_value(row, "Current Phase") or ""),
        ]
    )


def _row_pos_value(row: dict[str, Any]) -> float | None:
    for header in (
        "Source-Only PoS",
        "Source Only PoS",
        "Primary Source-Only PoS",
        "Primary Source Only PoS",
        "Phase LOA",
        "PoS",
        "Probability of Success",
    ):
        value = _get_value(row, header)
        if value is not None:
            return to_float(value)
    for key, value in row.items():
        normalized_key = norm(str(key))
        if "pos" in normalized_key and "weighted" not in normalized_key and value is not None:
            return to_float(value)
    return None


def _get_value(row: dict[str, Any], wanted_header: str) -> Any:
    wanted = norm(wanted_header)
    for key, value in row.items():
        if norm(str(key)) == wanted:
            return value
    return None


def _normalize_words(value: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", value.casefold())).strip()
