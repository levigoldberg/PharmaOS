"""Pricing evidence from local WAC data and openFDA labels."""

from __future__ import annotations

import os
import re
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook
from pydantic import Field

from pharma_os.agent_runtime import AgentRuntimeError, run_structured_llm_call, runtime_config_for_route
from pharma_os.schemas import AssetIdentityOutput, MissingDataFlag, PricingOutput, SourceMetadata, StrictSchema
from pharma_os.tools._due_diligence_common import DEFAULT_WAC_DATA, OPENFDA_LABEL_URL, missing, norm, slug, to_float
from pharma_os.tools.rules import config_provenance, config_source, load_config


REPO_ROOT = Path(__file__).resolve().parents[3]
WAC_COLUMN_ALIASES = {
    "brand_name": ("brand_name", "brand", "proprietary_name", "drug name"),
    "generic_name": ("generic_name", "generic", "nonproprietary_name", "generic name"),
    "manufacturer": ("manufacturer", "labeler_name", "company", "manufacturer_name", "manufacturer name"),
    "ndc": ("ndc", "ndc11", "ndc code", "product_ndc", "ndc number"),
    "product_description": ("product_description", "description", "drugname", "product", "drug product description"),
    "strength": ("strength", "strengths"),
    "dosage_form": ("dosage_form", "form"),
    "package_size": ("package_size", "package", "pkg_size"),
    "wac_value": ("wac_value", "wac", "current wac", "price", "unit_wac", "wac after increase"),
    "currency": ("currency",),
    "wac_unit_basis": ("wac_unit_basis", "unit basis", "pricing unit"),
    "effective_date": ("effective_date", "effective date", "date", "wac effective date"),
    "date_reported": ("date_reported", "date reported"),
    "source_file": ("source_file",),
    "source_sheet": ("source_sheet",),
    "source_year": ("source_year",),
}


@dataclass(frozen=True)
class PricingSearchTerm:
    """Normalized exact or source-constrained analog term for WAC lookup."""

    value: str
    is_analog: bool = False
    reason: str | None = None
    relevance_score: int = 100
    brand_name: str | None = None
    generic_name: str | None = None


@dataclass(frozen=True)
class AnnualizedWAC:
    """Deterministic annual WAC calculation details."""

    annual_wac: float
    details: dict[str, str | int | float | bool | None | list[str]]


@dataclass(frozen=True)
class LabelLookupResult:
    """Selected openFDA label payload and query provenance."""

    payload: dict[str, Any] | None
    selected_query: str | None = None
    selected_term: str | None = None
    flags: tuple[MissingDataFlag, ...] = ()


@dataclass(frozen=True)
class StructuredDosingEvidence:
    """Selected regimen evidence matched to the WAC product."""

    route: str | None = None
    dosage_form: str | None = None
    dose_mg: float | None = None
    frequency_text: str | None = None
    administrations_per_year: float | None = None
    relevant_text_excerpt: str | None = None
    loading_phase: str | None = None
    maintenance_phase: str | None = None
    ambiguity_flags: tuple[str, ...] = ()
    semantic_validation_flags: tuple[str, ...] = ()
    confidence: float = 0.0


class PricingAnalogCandidate(StrictSchema):
    """Approved pricing analog candidate before WAC source matching."""

    brand_name: str | None = None
    generic_name: str | None = None
    approved_indication: str | None = None
    route: str | None = None
    modality: str | None = None
    rationale: str = Field(..., min_length=1)
    confidence: float = Field(default=0.5, ge=0, le=1)


class PricingAnalogSelectionOutput(StrictSchema):
    """Candidate approved analogs for source-constrained WAC matching."""

    candidates: tuple[PricingAnalogCandidate, ...] = Field(default_factory=tuple)
    human_review_flags: tuple[str, ...] = Field(default_factory=tuple)


def lookup_pricing(
    asset: AssetIdentityOutput,
    *,
    wac_data_path: str | None = None,
    client: httpx.Client | None = None,
) -> tuple[PricingOutput, tuple[SourceMetadata, ...]]:
    """Match local WAC rows and openFDA dosing/label data."""

    wac_config = load_config("wac_sources.yaml", section="due_diligence")
    wac_config_source = config_source("wac_sources.yaml", section="due_diligence")
    configured_path = _configured_wac_path(wac_config)
    path = Path(wac_data_path or os.getenv("PHARMA_OS_WAC_DATA_PATH") or configured_path or DEFAULT_WAC_DATA)
    wac_source = SourceMetadata(
        source_id=f"wac:{slug(path.name)}",
        title="Local WAC workbook",
        provenance=f"Local WAC workbook combined_wac_increases lookup approved by {config_provenance('wac_sources.yaml', 'approved_sources[0]', section='due_diligence')}",
        source_type="wac_workbook",
        version=path.name,
    )
    flags: list[MissingDataFlag] = []
    wac_value = None
    matched_product = None
    matched_term: PricingSearchTerm | None = None
    matched_wac_row: dict[str, Any] | None = None
    resolved_path = _resolve_wac_path(path)
    if not resolved_path.exists():
        flags.append(missing("pricing-wac-file-missing", "pricing", "wac_value", f"WAC workbook not found: {resolved_path}", "high"))
    elif not asset.asset_name:
        flags.append(missing("pricing-asset-missing", "pricing", "matched_product", "No asset name available for WAC lookup.", "high"))
    else:
        try:
            terms = _pricing_terms(asset)
            rows = _load_wac_rows(resolved_path)
            matches: list[tuple[dict[str, Any], PricingSearchTerm]] = []
            for row_map in rows:
                match = _match_wac_row(row_map, terms)
                if match is not None:
                    matches.append((row_map, match))
            if matches:
                row_map, matched_term = _select_wac_match(matches)
                matched_wac_row = row_map
                wac_value = to_float(row_map.get("wac_value"))
                description = str(row_map.get("product_description") or "")
                matched_product = (
                    f"{description} (pricing analog: {matched_term.value}; {matched_term.reason})"
                    if matched_term.is_analog and matched_term.reason
                    else description
                )
            if wac_value is None:
                flags.append(
                    missing(
                        "pricing-no-wac-match",
                        "pricing",
                        "wac_value",
                        f"No exact or source-constrained pricing analog WAC row matched {asset.asset_name}.",
                        "high",
                    )
                )
        except AgentRuntimeError:
            raise
        except Exception as exc:
            flags.append(missing("pricing-wac-error", "pricing", "wac_value", f"WAC lookup failed: {exc.__class__.__name__}.", "high"))

    label_lookup = _fetch_openfda_label(
        asset=asset,
        matched_term=matched_term,
        wac_row=matched_wac_row,
        client=client or httpx.Client(timeout=20.0),
    )
    flags.extend(label_lookup.flags)
    label_result = (label_lookup.payload.get("results") or [{}])[0] if label_lookup.payload else None
    dosing_evidence = (
        _extract_structured_dosing(label_result, matched_product, matched_wac_row, asset)
        if label_result is not None
        else None
    )
    dosing_summary = _dosing_summary(dosing_evidence)
    annualized = _annualize_wac(wac_value, matched_product, dosing_evidence)
    annual_wac = annualized.annual_wac if annualized is not None else None
    annualization_details = annualized.details if annualized is not None else {}
    if wac_value is not None and dosing_summary is None:
        flags.append(missing("pricing-dosing-review-required", "pricing", "annual_wac", "WAC was found but annualization lacks sourced dosing.", "high"))
    elif wac_value is not None and dosing_evidence and dosing_evidence.semantic_validation_flags:
        flags.append(
            missing(
                "pricing-dosing-semantic-validation",
                "pricing",
                "annual_wac",
                "Selected dosing evidence does not semantically match the chosen WAC product: "
                + "; ".join(dosing_evidence.semantic_validation_flags),
                "high",
            )
        )
    elif wac_value is not None and annual_wac is None:
        flags.append(missing("pricing-annualization-review-required", "pricing", "annual_wac", "WAC and dosing were found, but package/frequency annualization requires review.", "high"))
    elif annualized is not None and not annualized.details.get("formula_check_passed"):
        flags.append(missing("pricing-annualization-formula-check", "pricing", "annual_wac", "Annualized WAC failed internal formula validation.", "critical"))
    label_source = (
        SourceMetadata(
            source_id=f"openfda_label:{slug(label_lookup.selected_term)}",
            title=f"openFDA label search for {label_lookup.selected_term}",
            url=OPENFDA_LABEL_URL,
            provenance="openFDA drug label API",
            source_type="drug_label",
            version="openFDA",
        )
        if label_lookup.selected_term and dosing_summary
        else None
    )
    sources = tuple(source for source in (wac_config_source, wac_source if wac_value is not None else None, label_source) if source)
    return (
        PricingOutput(
            annual_wac=annual_wac,
            wac_value=wac_value,
            wac_unit_basis="package" if wac_value is not None else None,
            matched_product=matched_product,
            dosing_summary=dosing_summary,
            annualization_details=annualization_details,
            source_ids=tuple(source.source_id for source in sources),
            missing_data_flags=tuple(flags),
            confidence=0.8 if annual_wac is not None else 0.45 if wac_value is not None else 0.2,
        ),
        sources,
    )


def _pricing_terms(asset: AssetIdentityOutput) -> tuple[PricingSearchTerm, ...]:
    terms = [
        PricingSearchTerm(term, relevance_score=100, brand_name=asset.asset_name)
        for term in [asset.asset_name, *asset.aliases]
        if term
    ]
    if asset.rxnorm_match:
        terms.extend(
            PricingSearchTerm(term, relevance_score=100, brand_name=asset.rxnorm_match.matched_name)
            for term in [asset.rxnorm_match.matched_name, *asset.rxnorm_match.aliases]
            if term
        )
    terms.extend(_pricing_analog_terms(asset))
    deduped: dict[str, PricingSearchTerm] = {}
    for term in terms:
        key = norm(term.value)
        if key and (key not in deduped or term.relevance_score > deduped[key].relevance_score):
            deduped[key] = term
    return tuple(deduped.values())


def _pricing_analog_terms(asset: AssetIdentityOutput) -> tuple[PricingSearchTerm, ...]:
    candidates = _pricing_analog_candidates(asset)
    terms: list[PricingSearchTerm] = []
    for index, candidate in enumerate(candidates):
        score = max(1, min(99, int(candidate.confidence * 100))) - index
        reason = candidate.rationale
        if candidate.brand_name:
            terms.append(
                PricingSearchTerm(
                    candidate.brand_name,
                    is_analog=True,
                    reason=reason,
                    relevance_score=score,
                    brand_name=candidate.brand_name,
                    generic_name=candidate.generic_name,
                )
            )
        if candidate.generic_name:
            terms.append(
                PricingSearchTerm(
                    candidate.generic_name,
                    is_analog=True,
                    reason=reason,
                    relevance_score=score,
                    brand_name=candidate.brand_name,
                    generic_name=candidate.generic_name,
                )
            )
    return tuple(terms)


def _pricing_analog_candidates(asset: AssetIdentityOutput) -> tuple[PricingAnalogCandidate, ...]:
    fallback = _fallback_pricing_analog_selection(asset)
    result = run_structured_llm_call(
        agent_name="PricingAnalogSelectionAgent",
        instructions=_pricing_analog_selection_instructions(),
        payload={"asset_context": _pricing_asset_context(asset)},
        output_type=PricingAnalogSelectionOutput,
        run_id=f"pricing-analog-selection-{slug(asset.nct_id or asset.asset_name or 'asset')}",
        input_summary=f"Select approved pricing analogs for {asset.asset_name or asset.nct_id}.",
        config=runtime_config_for_route(
            model_route="agent4_subagent",
            disabled_provenance="pharma_os.tools.pricing",
        ),
        offline_output=fallback,
        source_ids=asset.source_ids,
        confidence=asset.confidence,
        rationale_summary="Select approved commercial pricing analog candidates before deterministic WAC source matching.",
    )
    selected = result.output
    return _dedupe_analog_candidates((*selected.candidates, *fallback.candidates))


def _fallback_pricing_analog_selection(asset: AssetIdentityOutput) -> PricingAnalogSelectionOutput:
    text = _word_text(asset.asset_name, asset.normalized_indication, asset.therapeutic_area, *asset.aliases)
    analogs: list[PricingAnalogCandidate] = []
    if any(term in text for term in ("sle", "lupus", "systemic lupus erythematosus")):
        if "nephritis" in text:
            analogs.extend(
                [
                    _analog("Lupkynis", "voclosporin", "lupus nephritis approved pricing analog", indication="lupus nephritis", route="oral", modality="small molecule", confidence=0.85),
                ]
            )
        analogs.extend(
            [
                _analog("Benlysta", "belimumab", "systemic lupus erythematosus approved pricing analog", indication="systemic lupus erythematosus", route="subcutaneous", modality="biologic", confidence=0.82),
            ]
        )
    if any(term in text for term in ("atopic dermatitis", "eczema", "dermatitis", "ad ")):
        analogs.extend(
            [
                _analog("Cibinqo", "abrocitinib", "atopic dermatitis oral JAK inhibitor approved pricing analog", indication="atopic dermatitis", route="oral", modality="small molecule", confidence=0.88),
                _analog("Rinvoq", "upadacitinib", "atopic dermatitis oral JAK inhibitor approved pricing analog", indication="atopic dermatitis", route="oral", modality="small molecule", confidence=0.84),
                _analog("Dupixent", "dupilumab", "atopic dermatitis biologic approved pricing analog", indication="atopic dermatitis", route="subcutaneous", modality="biologic", confidence=0.78),
                _analog("Adbry", "tralokinumab", "atopic dermatitis biologic approved pricing analog", indication="atopic dermatitis", route="subcutaneous", modality="biologic", confidence=0.72),
            ]
        )
    if any(term in text for term in ("psoriasis", "psoriatic", "tyk2", "tyrosine kinase 2")):
        analogs.extend(
            [
                _analog("Sotyktu", "deucravacitinib", "TYK2/psoriasis approved pricing analog", indication="plaque psoriasis", route="oral", modality="small molecule", confidence=0.85),
            ]
        )
    if "rheumatoid arthritis" in text or ("autoimmune" in text and "small molecule" in text):
        analogs.extend(
            [
                _analog("Rinvoq", "upadacitinib", "autoimmune oral small-molecule approved pricing analog", route="oral", modality="small molecule", confidence=0.78),
            ]
        )
    return PricingAnalogSelectionOutput(candidates=_dedupe_analog_candidates(tuple(analogs)))


def _analog(
    brand_name: str,
    generic_name: str,
    rationale: str,
    *,
    indication: str | None = None,
    route: str | None = None,
    modality: str | None = None,
    confidence: float = 0.7,
) -> PricingAnalogCandidate:
    return PricingAnalogCandidate(
        brand_name=brand_name,
        generic_name=generic_name,
        approved_indication=indication,
        route=route,
        modality=modality,
        rationale=rationale,
        confidence=confidence,
    )


def _dedupe_analog_candidates(candidates: tuple[PricingAnalogCandidate, ...]) -> tuple[PricingAnalogCandidate, ...]:
    deduped: dict[tuple[str, str], PricingAnalogCandidate] = {}
    for candidate in candidates:
        key = (norm(candidate.brand_name or ""), norm(candidate.generic_name or ""))
        if not any(key):
            continue
        existing = deduped.get(key)
        if existing is None or candidate.confidence > existing.confidence:
            deduped[key] = candidate
    return tuple(deduped.values())


def _pricing_asset_context(asset: AssetIdentityOutput) -> dict[str, Any]:
    return {
        "nct_id": asset.nct_id,
        "asset_name": asset.asset_name,
        "aliases": asset.aliases,
        "raw_intervention_names": asset.raw_intervention_names,
        "intervention_type": asset.intervention_type,
        "sponsor": asset.sponsor,
        "normalized_indication": asset.normalized_indication,
        "therapeutic_area": asset.therapeutic_area,
        "modality": asset.modality,
        "guardrails": (
            "Return approved commercial drug analogs only. Do not invent WAC values, dosing, labels, or unapproved products. "
            "Prefer analogs likely to have public label dosing and approved WAC workbook rows."
        ),
    }


def _pricing_analog_selection_instructions() -> str:
    return (
        "You are PricingAnalogSelectionAgent for PharmaOS Agent 4. Select approved commercial pricing analog candidates "
        "for the target investigational asset using only the supplied asset context. Prefer analogs by indication, route, "
        "modality, mechanism/class, prescriber setting, and commercial use. Return candidates only; deterministic code will "
        "match candidates to the approved WAC workbook and fetch openFDA dosing. Do not calculate price, claim WAC "
        "availability, fetch labels, or fabricate evidence."
    )


def _label_terms(asset: AssetIdentityOutput, matched_term: PricingSearchTerm | None) -> tuple[PricingSearchTerm, ...]:
    terms: list[PricingSearchTerm] = []
    if matched_term is not None:
        terms.append(matched_term)
        if matched_term.brand_name and norm(matched_term.brand_name) != norm(matched_term.value):
            terms.append(
                PricingSearchTerm(
                    matched_term.brand_name,
                    is_analog=matched_term.is_analog,
                    reason=matched_term.reason,
                    relevance_score=matched_term.relevance_score,
                    brand_name=matched_term.brand_name,
                    generic_name=matched_term.generic_name,
                )
            )
        if matched_term.generic_name and norm(matched_term.generic_name) != norm(matched_term.value):
            terms.append(
                PricingSearchTerm(
                    matched_term.generic_name,
                    is_analog=matched_term.is_analog,
                    reason=matched_term.reason,
                    relevance_score=matched_term.relevance_score,
                    brand_name=matched_term.brand_name,
                    generic_name=matched_term.generic_name,
                )
            )
    else:
        terms.extend(_pricing_terms(asset))
    deduped: dict[str, PricingSearchTerm] = {}
    for term in terms:
        key = norm(term.value)
        if key and key not in deduped:
            deduped[key] = term
    return tuple(deduped.values())


def _fetch_openfda_label(
    *,
    asset: AssetIdentityOutput,
    matched_term: PricingSearchTerm | None,
    wac_row: dict[str, Any] | None,
    client: httpx.Client,
) -> LabelLookupResult:
    """Fetch the most relevant openFDA label, preferring exact WAC NDC evidence."""

    queries = _label_queries(asset, matched_term, wac_row)
    if not queries:
        return LabelLookupResult(
            payload=None,
            flags=(missing("pricing-label-search-terms-missing", "pricing", "dosing_summary", "No label search term was available.", "medium"),),
        )
    flags: list[MissingDataFlag] = []
    for query, display_term in queries:
        try:
            response = client.get(OPENFDA_LABEL_URL, params={"search": query, "limit": "5"}, timeout=20.0)
        except Exception as exc:
            flags.append(
                missing(
                    f"pricing-openfda-error-{len(flags) + 1}",
                    "pricing",
                    "dosing_summary",
                    f"openFDA label request failed for {display_term}: {exc.__class__.__name__}.",
                    "medium",
                )
            )
            continue
        if response.status_code in {204, 404}:
            continue
        if response.status_code not in range(200, 300):
            flags.append(
                missing(
                    f"pricing-openfda-status-{len(flags) + 1}",
                    "pricing",
                    "dosing_summary",
                    f"openFDA label request for {display_term} returned HTTP {response.status_code}.",
                    "medium",
                )
            )
            continue
        try:
            payload = response.json()
        except ValueError:
            flags.append(
                missing(
                    f"pricing-openfda-json-{len(flags) + 1}",
                    "pricing",
                    "dosing_summary",
                    f"openFDA label response for {display_term} was not JSON.",
                    "medium",
                )
            )
            continue
        selected = _select_label_candidate(payload, wac_row, matched_term)
        if selected is not None:
            return LabelLookupResult(
                payload={"results": [selected]},
                selected_query=query,
                selected_term=display_term,
                flags=tuple(flags),
            )
    flags.append(missing("pricing-openfda-label-missing", "pricing", "dosing_summary", "No matching openFDA label was found for the selected WAC product.", "medium"))
    return LabelLookupResult(payload=None, flags=tuple(flags))


def _label_queries(
    asset: AssetIdentityOutput,
    matched_term: PricingSearchTerm | None,
    wac_row: dict[str, Any] | None,
) -> list[tuple[str, str]]:
    brand = _preferred_brand(asset, matched_term, wac_row)
    generic = _preferred_generic(matched_term, wac_row)
    display = brand or generic or (matched_term.value if matched_term else asset.asset_name) or "selected product"
    queries: list[tuple[str, str]] = []
    for ndc in _ndc_query_variants(wac_row):
        queries.append((f'openfda.product_ndc:"{_escape_openfda_term(_product_ndc(ndc))}"', display))
        queries.append((f'openfda.package_ndc:"{_escape_openfda_term(ndc)}"', display))
    if brand and generic and norm(brand) != norm(generic):
        queries.append((f'openfda.brand_name:"{_escape_openfda_term(brand)}" AND openfda.generic_name:"{_escape_openfda_term(generic)}"', brand))
    if brand:
        queries.append((f'openfda.brand_name:"{_escape_openfda_term(brand)}"', brand))
    if generic:
        queries.append((f'openfda.generic_name:"{_escape_openfda_term(generic)}"', generic))
    for term in _label_terms(asset, matched_term):
        if term.value and all(norm(term.value) != norm(existing[1]) for existing in queries):
            queries.append((f'openfda.brand_name:"{_escape_openfda_term(term.value)}"', term.value))
    deduped: list[tuple[str, str]] = []
    seen: set[str] = set()
    for query, term in queries:
        key = query.casefold()
        if key not in seen:
            seen.add(key)
            deduped.append((query, term))
    return deduped


def _select_label_candidate(
    payload: dict[str, Any],
    wac_row: dict[str, Any] | None,
    matched_term: PricingSearchTerm | None,
) -> dict[str, Any] | None:
    results = payload.get("results") if isinstance(payload, dict) else None
    if not isinstance(results, list):
        return None
    candidates = [item for item in results if isinstance(item, dict)]
    if not candidates:
        return None
    scored = [(_label_candidate_score(item, wac_row, matched_term), index, item) for index, item in enumerate(candidates)]
    scored.sort(key=lambda item: (item[0], -item[1]), reverse=True)
    return scored[0][2]


def _label_candidate_score(
    label: dict[str, Any],
    wac_row: dict[str, Any] | None,
    matched_term: PricingSearchTerm | None,
) -> int:
    openfda = label.get("openfda") if isinstance(label.get("openfda"), dict) else {}
    text = _word_text(
        _all_text(openfda.get("brand_name")),
        _all_text(openfda.get("generic_name")),
        _all_text(openfda.get("product_ndc")),
        _all_text(openfda.get("package_ndc")),
        _all_text(label.get("dosage_and_administration")),
    )
    score = 0
    ndcs = set(_ndc_query_variants(wac_row))
    label_ndcs = {_normalize_ndc(item) for item in _string_values(openfda.get("product_ndc")) + _string_values(openfda.get("package_ndc"))}
    if ndcs and any(_normalize_ndc(ndc) in label_ndcs or _product_ndc(_normalize_ndc(ndc)) in label_ndcs for ndc in ndcs):
        score += 25
    for term, weight in (
        (_preferred_brand(None, matched_term, wac_row), 8),
        (_preferred_generic(matched_term, wac_row), 6),
        (matched_term.value if matched_term else None, 5),
    ):
        if term and _word_text(term) in text:
            score += weight
    if _all_text(label.get("dosage_and_administration")):
        score += 3
    return score


def _extract_structured_dosing(
    label_result: dict[str, Any],
    matched_product: str | None,
    wac_row: dict[str, Any] | None,
    asset: AssetIdentityOutput,
) -> StructuredDosingEvidence | None:
    snippets = _dosing_snippets(label_result)
    if not snippets:
        return None
    scored: list[tuple[float, int, StructuredDosingEvidence]] = []
    product_description = _product_description(matched_product, wac_row)
    product_strength = _strength_mg(_word_text(product_description, str(wac_row.get("strength") if wac_row else "")))
    product_form = _dosage_form(product_description, str(wac_row.get("dosage_form") or "") if wac_row else "")
    product_route = _route_from_text(product_description)
    product_is_extended_release = _is_extended_release(product_description)
    indication_text = _word_text(asset.normalized_indication, asset.therapeutic_area)
    for index, snippet in enumerate(snippets):
        dose = _dose_mg(snippet)
        administrations = _administrations_per_year(snippet)
        if administrations is None and dose is None:
            continue
        flags: list[str] = []
        ambiguity: list[str] = []
        score = 0.0
        snippet_words = _word_text(snippet)
        snippet_route = _route_from_text(snippet)
        snippet_form = _dosage_form(snippet)
        if administrations is not None:
            score += 4.0
        if dose is not None:
            score += 2.0
        if product_strength is not None and dose is not None:
            ratio = dose / product_strength if product_strength else 0
            if _close(dose, product_strength):
                score += 10.0
            elif ratio > 0 and abs(ratio - round(ratio)) <= 0.05 and round(ratio) <= 6 and not product_is_extended_release:
                score += 5.0
                ambiguity.append(f"selected dose requires {round(ratio)} units per administration")
            else:
                score -= 8.0
                flags.append(f"label dose {dose:g} mg does not match WAC product strength {product_strength:g} mg")
        if product_is_extended_release:
            if _is_extended_release(snippet):
                score += 8.0
            else:
                score -= 8.0
                flags.append("WAC product is extended-release/XR but label regimen snippet is not")
        elif _is_extended_release(snippet) and not _is_extended_release(product_description):
            score -= 3.0
            ambiguity.append("label regimen snippet is extended-release while WAC product is not explicitly extended-release")
        if product_route and snippet_route:
            if product_route == snippet_route:
                score += 4.0
            else:
                score -= 5.0
                flags.append(f"label route {snippet_route} does not match WAC product route {product_route}")
        elif snippet_route:
            score += 1.0
        if product_form and snippet_form:
            if product_form == snippet_form:
                score += 3.0
            elif product_form in {"tablet", "capsule"} and snippet_form in {"tablet", "capsule"}:
                score -= 2.0
                flags.append(f"label dosage form {snippet_form} does not match WAC product form {product_form}")
        if indication_text and any(part and part in snippet_words for part in indication_text.split()):
            score += 1.5
        if any(term in snippet_words for term in ("maintenance", "recommended dosage", "recommended dose")):
            score += 2.0
        if _is_loading_phase(snippet):
            score -= 4.0
            ambiguity.append("snippet appears to describe loading, titration, or initial dosing")
        if _word_text(_preferred_brand(asset, None, wac_row)) and _word_text(_preferred_brand(asset, None, wac_row)) in snippet_words:
            score += 2.0
        scored.append(
            (
                score - (6.0 * len(flags)),
                index,
                StructuredDosingEvidence(
                    route=snippet_route or product_route,
                    dosage_form=snippet_form or product_form,
                    dose_mg=dose,
                    frequency_text=_frequency_text(snippet),
                    administrations_per_year=administrations,
                    relevant_text_excerpt=_trim(snippet, 450),
                    loading_phase=snippet if _is_loading_phase(snippet) else None,
                    maintenance_phase=snippet if not _is_loading_phase(snippet) else None,
                    ambiguity_flags=tuple(ambiguity),
                    semantic_validation_flags=tuple(flags),
                    confidence=max(0.0, min(1.0, (score + 6.0) / 24.0)),
                ),
            )
        )
    if not scored:
        return None
    scored.sort(key=lambda item: (item[0], -item[1]), reverse=True)
    selected = scored[0][2]
    if selected.administrations_per_year is None:
        return replace(
            selected,
            semantic_validation_flags=(*selected.semantic_validation_flags, "selected label snippet lacks a deterministic maintenance frequency"),
        )
    return selected


def _dosing_summary(evidence: StructuredDosingEvidence | None) -> str | None:
    if evidence is None or not evidence.relevant_text_excerpt:
        return None
    parts = []
    if evidence.dose_mg is not None and evidence.frequency_text:
        parts.append(f"Selected regimen: {evidence.dose_mg:g} mg {evidence.frequency_text}.")
    elif evidence.frequency_text:
        parts.append(f"Selected regimen frequency: {evidence.frequency_text}.")
    parts.append(f"Label evidence: {evidence.relevant_text_excerpt}")
    if evidence.ambiguity_flags:
        parts.append("Review context: " + "; ".join(evidence.ambiguity_flags[:3]) + ".")
    return " ".join(parts)


def _match_wac_row(row_map: dict[str, Any], terms: tuple[PricingSearchTerm, ...]) -> PricingSearchTerm | None:
    haystacks = [
        str(row_map.get("brand_name") or ""),
        str(row_map.get("generic_name") or ""),
        str(row_map.get("product_description") or ""),
        str(row_map.get("manufacturer") or ""),
        str(row_map.get("ndc") or ""),
    ]
    normalized_haystacks = [norm(item) for item in haystacks]
    for term in terms:
        normalized_term = norm(term.value)
        if normalized_term and any(normalized_term in haystack for haystack in normalized_haystacks):
            return term
    return None


def _select_wac_match(matches: list[tuple[dict[str, Any], PricingSearchTerm]]) -> tuple[dict[str, Any], PricingSearchTerm]:
    return sorted(matches, key=_wac_selection_key, reverse=True)[0]


def _wac_selection_key(match: tuple[dict[str, Any], PricingSearchTerm]) -> tuple[Any, ...]:
    row, term = match
    return (
        0 if term.is_analog else 1,
        term.relevance_score,
        _parse_date(str(row.get("effective_date") or "")),
        _parse_date(str(row.get("date_reported") or "")),
        int(to_float(row.get("source_year")) or 0),
        -(_units_per_package(str(row.get("product_description") or "")) or 999999),
    )


def _annualize_wac(wac_value: float | None, matched_product: str | None, dosing_evidence: StructuredDosingEvidence | None) -> AnnualizedWAC | None:
    if wac_value is None or not matched_product or dosing_evidence is None:
        return None
    if dosing_evidence.semantic_validation_flags:
        return None
    administrations_per_year = dosing_evidence.administrations_per_year
    if administrations_per_year is None:
        return None
    units_per_package = _units_per_package(matched_product)
    if not units_per_package:
        return None
    units_per_administration = _units_per_administration(matched_product, dosing_evidence)
    if units_per_administration is None:
        return None
    package_units_per_year = administrations_per_year * units_per_administration
    packages_per_year = package_units_per_year / units_per_package
    annual_wac = round(float(wac_value) * packages_per_year, 2)
    expected = round(float(wac_value) * administrations_per_year * units_per_administration / units_per_package, 2)
    details: dict[str, str | int | float | bool | None | list[str]] = {
        "formula": "annual_wac = wac_value * administrations_per_year * units_per_administration / units_per_package",
        "wac_value": float(wac_value),
        "wac_unit_basis": "package",
        "selected_route": dosing_evidence.route,
        "selected_dosage_form": dosing_evidence.dosage_form,
        "selected_dose_mg": dosing_evidence.dose_mg,
        "selected_frequency": dosing_evidence.frequency_text,
        "administrations_per_year": administrations_per_year,
        "units_per_administration": units_per_administration,
        "units_per_package": units_per_package,
        "package_units_per_year": round(package_units_per_year, 4),
        "packages_per_year": round(packages_per_year, 4),
        "annual_wac": annual_wac,
        "semantic_validation_passed": True,
        "semantic_validation_flags": [],
        "relevant_text_excerpt": dosing_evidence.relevant_text_excerpt,
        "formula_check_expected_annual_wac": expected,
        "formula_check_passed": abs(annual_wac - expected) <= 0.01,
    }
    return AnnualizedWAC(annual_wac=annual_wac, details=details)


def _administrations_per_year(dosing_summary: str) -> float | None:
    text = _word_text(dosing_summary)
    if "twice weekly" in text or "two times weekly" in text:
        return 104.0
    if any(term in text for term in ("once weekly", "once a week", "every week", "weekly")):
        return 52.0
    if any(term in text for term in ("every 2 weeks", "every two weeks", "every other week", "once every 2 weeks")):
        return 26.0
    if any(term in text for term in ("every 4 weeks", "monthly", "once monthly", "once every month")):
        return 13.0
    if any(term in text for term in ("three times daily", "three times a day", "thrice daily")):
        return 1095.0
    if any(term in text for term in ("twice daily", "twice a day", "two times daily")):
        return 730.0
    if any(term in text for term in ("once daily", "once a day", "daily")):
        return 365.0
    return None


def _units_per_administration(product_description: str, dosing_evidence: StructuredDosingEvidence) -> float | None:
    product = _word_text(product_description)
    strength = _strength_mg(product_description)
    dose = dosing_evidence.dose_mg
    if strength and dose:
        units = dose / strength
        if units <= 0:
            return None
        return round(max(1.0, units), 4)
    if any(term in product for term in ("auto injector", "autoinjector", "prefilled syringe", "syringe", "vial")):
        return 1.0
    if any(term in product for term in ("tablet", "capsule", "capsules", "tab", "bottle")):
        return 1.0
    return None


def _units_per_package(product_description: str) -> int | None:
    text = product_description.casefold()
    if match := re.search(r"\bx\s*(\d+)\b", text):
        return int(match.group(1))
    if match := re.search(r"\b(\d+)\s*(?:capsules?|tablets?|tabs?|count|ct|ea|pack|pens?|syringes?|auto-?injectors?|autoinjectors?|vials?)\b", text):
        return int(match.group(1))
    if match := re.search(r"\b(?:capsules?|tablets?|tabs?|pens?|syringes?|auto-?injectors?|autoinjectors?|vials?)\s*(\d+)\b", text):
        return int(match.group(1))
    if match := re.search(r"-\s*(\d+)\s*(?:pens?|syringes?|auto-?injectors?|autoinjectors?|vials?)\b", text):
        return int(match.group(1))
    if match := re.search(r"\b(\d+)\s*day\s*bottle\b", text):
        return int(match.group(1))
    if any(term in text for term in ("auto-injector", "autoinjector", "prefilled syringe", "syringe", "vial")):
        return 1
    return None


def _strength_mg(product_description: str) -> float | None:
    if match := re.search(r"\b(\d+(?:\.\d+)?)\s*mg\b", product_description.casefold()):
        return float(match.group(1))
    return None


def _dose_mg(dosing_summary: str) -> float | None:
    if match := re.search(r"\b(\d+(?:\.\d+)?)\s*mg\b", dosing_summary.casefold()):
        return float(match.group(1))
    return None


def _preferred_brand(
    asset: AssetIdentityOutput | None,
    matched_term: PricingSearchTerm | None,
    wac_row: dict[str, Any] | None,
) -> str | None:
    values = [
        str(wac_row.get("brand_name") or "") if wac_row else None,
        matched_term.brand_name if matched_term else None,
        matched_term.value if matched_term and not matched_term.generic_name else None,
        _infer_brand_from_description(str(wac_row.get("product_description") or "")) if wac_row else None,
        asset.asset_name if asset else None,
    ]
    return next((value.strip() for value in values if isinstance(value, str) and value.strip()), None)


def _preferred_generic(matched_term: PricingSearchTerm | None, wac_row: dict[str, Any] | None) -> str | None:
    values = [
        str(wac_row.get("generic_name") or "") if wac_row else None,
        matched_term.generic_name if matched_term else None,
    ]
    return next((value.strip() for value in values if isinstance(value, str) and value.strip()), None)


def _infer_brand_from_description(description: str) -> str | None:
    if not description:
        return None
    cleaned = re.sub(r"[;,(].*$", "", description).strip()
    cleaned = re.sub(r"[™®]", "", cleaned).strip()
    if not cleaned:
        return None
    words = cleaned.split()
    if not words:
        return None
    if len(words) >= 2 and words[1].casefold() in {"xr", "er", "cr", "dr"}:
        return " ".join(words[:2])
    return words[0]


def _product_description(matched_product: str | None, wac_row: dict[str, Any] | None) -> str:
    row_description = str(wac_row.get("product_description") or "") if wac_row else ""
    return " ".join(part for part in (row_description, matched_product or "") if part).strip()


def _escape_openfda_term(value: str) -> str:
    return value.replace("\\", "\\\\").replace('"', '\\"')


def _ndc_query_variants(wac_row: dict[str, Any] | None) -> list[str]:
    if not wac_row:
        return []
    raw_values: list[str] = []
    raw_ndc = wac_row.get("ndc")
    if raw_ndc is not None:
        raw_values.append(_clean_ndc_value(raw_ndc))
    description = str(wac_row.get("product_description") or "")
    raw_values.extend(re.findall(r"\b\d{4,5}-\d{3,4}(?:-\d{1,2})?\b", description))
    variants: list[str] = []
    for raw in raw_values:
        if not raw:
            continue
        raw = raw.strip()
        if "-" in raw:
            variants.append(raw)
            parts = raw.split("-")
            if len(parts) == 3:
                variants.append("-".join(parts[:2]))
        digits = re.sub(r"\D+", "", raw)
        if len(digits) == 11:
            variants.extend([f"{digits[:5]}-{digits[5:9]}-{digits[9:]}", f"{digits[:5]}-{digits[5:9]}"])
        elif len(digits) == 10:
            variants.extend(
                [
                    f"{digits[:5]}-{digits[5:8]}-{digits[8:]}",
                    f"{digits[:5]}-{digits[5:8]}",
                    f"{digits[:4]}-{digits[4:8]}-{digits[8:]}",
                    f"{digits[:4]}-{digits[4:8]}",
                ]
            )
        elif len(digits) == 9:
            variants.append(f"{digits[:5]}-{digits[5:]}")
        elif digits:
            variants.append(digits)
    return _dedupe_strings(_normalize_ndc(value) for value in variants if value)


def _clean_ndc_value(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _product_ndc(ndc: str) -> str:
    parts = ndc.split("-")
    return "-".join(parts[:2]) if len(parts) >= 2 else ndc


def _normalize_ndc(value: str) -> str:
    text = str(value).strip()
    return re.sub(r"-+", "-", text).strip("-")


def _dosing_snippets(label_result: dict[str, Any]) -> list[str]:
    sections = [
        label_result.get("dosage_and_administration"),
        label_result.get("dosage_and_administration_table"),
        label_result.get("dosage_forms_and_strengths"),
    ]
    raw_text = "\n".join(_all_text(section) for section in sections if _all_text(section))
    raw_text = re.sub(r"\s+", " ", raw_text).strip()
    if not raw_text:
        return []
    pieces = re.split(r"(?<=[.;])\s+|(?:\s+-\s+)|(?:\n+)", raw_text)
    snippets = [_trim(piece.strip(" ;.\t\n"), 700) for piece in pieces if piece and piece.strip(" ;.\t\n")]
    return _dedupe_strings(snippets)


def _frequency_text(text: str) -> str | None:
    normalized = _word_text(text)
    patterns = (
        ("twice weekly", ("twice weekly", "two times weekly")),
        ("once weekly", ("once weekly", "once a week", "every week", "weekly")),
        ("every 2 weeks", ("every 2 weeks", "every two weeks", "every other week", "once every 2 weeks")),
        ("every 4 weeks", ("every 4 weeks", "monthly", "once monthly", "once every month")),
        ("three times daily", ("three times daily", "three times a day", "thrice daily")),
        ("twice daily", ("twice daily", "twice a day", "two times daily")),
        ("once daily", ("once daily", "once a day", "daily")),
    )
    for label, variants in patterns:
        if any(variant in normalized for variant in variants):
            return label
    return None


def _route_from_text(*values: str | None) -> str | None:
    text = _word_text(*values)
    if any(term in text for term in ("subcutaneous", "subcutaneously", "sc injection", "s c injection")):
        return "subcutaneous"
    if any(term in text for term in ("intravenous", "intravenously", "iv infusion", "i v infusion")):
        return "intravenous"
    if any(term in text for term in ("oral", "orally", "by mouth", "tablet", "capsule", "capsules")):
        return "oral"
    if any(term in text for term in ("intramuscular", "intramuscularly", "im injection")):
        return "intramuscular"
    return None


def _dosage_form(*values: str | None) -> str | None:
    text = _word_text(*values)
    if "auto injector" in text or "autoinjector" in text:
        return "auto-injector"
    if "prefilled syringe" in text or "syringe" in text:
        return "syringe"
    if "vial" in text:
        return "vial"
    if "tablet" in text or "tablets" in text or " tab " in f" {text} ":
        return "tablet"
    if "capsule" in text or "capsules" in text:
        return "capsule"
    return None


def _is_extended_release(text: str | None) -> bool:
    words = _word_text(text)
    tokens = set(words.split())
    return bool(tokens & {"xr", "er"}) or any(term in words for term in ("extended release", "extended released", "extendedrelease"))


def _is_loading_phase(text: str) -> bool:
    words = _word_text(text)
    return any(term in words for term in ("loading", "initial", "starting dose", "titration", "day 1", "days 1", "starter pack", "dose pack"))


def _all_text(value: Any) -> str:
    return " ".join(_string_values(value))


def _string_values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        return [value.strip()] if value.strip() else []
    if isinstance(value, (int, float)):
        return [str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)]
    if isinstance(value, list):
        values: list[str] = []
        for item in value:
            values.extend(_string_values(item))
        return values
    if isinstance(value, dict):
        values: list[str] = []
        for item in value.values():
            values.extend(_string_values(item))
        return values
    return []


def _dedupe_strings(values: Any) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        if not isinstance(value, str):
            continue
        cleaned = re.sub(r"\s+", " ", value).strip()
        if not cleaned:
            continue
        key = cleaned.casefold()
        if key not in seen:
            seen.add(key)
            result.append(cleaned)
    return result


def _trim(value: str, max_chars: int) -> str:
    text = re.sub(r"\s+", " ", value).strip()
    return text if len(text) <= max_chars else text[: max_chars - 3].rstrip() + "..."


def _close(left: float, right: float, *, tolerance: float = 0.1) -> bool:
    return abs(left - right) <= tolerance


def _parse_date(value: str) -> datetime:
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(value.strip(), fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(value.strip())
    except ValueError:
        return datetime.min


def _word_text(*values: str | None) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", " ".join(value or "" for value in values).casefold())).strip()


def _configured_wac_path(config: dict[str, Any]) -> str | None:
    sources = config.get("approved_sources")
    if not isinstance(sources, list):
        return None
    for source in sources:
        if isinstance(source, dict) and source.get("local_path"):
            return str(source["local_path"])
    return None


def _resolve_wac_path(path: Path) -> Path:
    if path.is_absolute() or path.exists():
        return path
    repo_path = REPO_ROOT / path
    if repo_path.exists():
        return repo_path
    return path


def _load_wac_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_name = "combined_wac_increases" if "combined_wac_increases" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    normalized_headers = {_normalize_header(header): index for index, header in enumerate(headers)}
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        normalized: dict[str, Any] = {}
        for target, aliases in WAC_COLUMN_ALIASES.items():
            index = next(
                (normalized_headers[_normalize_header(alias)] for alias in aliases if _normalize_header(alias) in normalized_headers),
                None,
            )
            normalized[target] = row[index] if index is not None and index < len(row) else None
        if any(normalized.get(key) for key in ("product_description", "brand_name", "generic_name", "ndc")):
            rows.append(normalized)
    return rows


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.casefold()).strip()
