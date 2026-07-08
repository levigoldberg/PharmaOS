"""Deterministic due-diligence tools adapted for PharmaOS."""

from __future__ import annotations

import math
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

from pharma_os.schemas import (
    AssumptionRecord,
    AssetIdentityOutput,
    ClinicalTrialRecord,
    CommercialModelOutput,
    MissingDataFlag,
    PatentCandidate,
    PatentExclusivityOutput,
    PoSOutput,
    PricingOutput,
    RevenueForecastYear,
    RNPVOutput,
    SourceMetadata,
)
from pharma_os.tools.rxnorm import RxNormClient, RxNormError
from pharma_os.tools.rules import config_provenance, config_source, config_source_id, human_override, load_config, load_rule_config


DEFAULT_POS_WORKBOOK = Path("data/Source_Based_PoS_Workbook.xlsx")
DEFAULT_WAC_DATA = Path("data/california_wac_data.xlsx")
OPENFDA_LABEL_URL = "https://api.fda.gov/drug/label.json"
LENS_PATENT_URL = "https://api.lens.org/patent/search"


def resolve_asset_identity(
    trial: ClinicalTrialRecord,
    *,
    rxnorm_client: RxNormClient | None = None,
) -> tuple[AssetIdentityOutput, tuple[SourceMetadata, ...]]:
    """Resolve asset identity from a normalized CT.gov trial and RxNorm."""

    sources: list[SourceMetadata] = [
        SourceMetadata(
            source_id=trial.source_id,
            title=trial.brief_title or trial.official_title or trial.nct_id,
            url=f"https://clinicaltrials.gov/study/{trial.nct_id}",
            authors=tuple(
                sponsor.name
                for sponsor in (trial.lead_sponsor, *trial.collaborators)
                if sponsor is not None
            ),
            provenance="ClinicalTrials.gov API v2 protocolSection",
            source_type="clinical_trial_registry",
            version="v2",
        )
    ]
    flags: list[MissingDataFlag] = []
    overrides = human_override(trial.nct_id)
    candidates = [
        item
        for item in trial.interventions
        if (item.type or "").upper() in {"DRUG", "BIOLOGICAL", "GENETIC"} and "placebo" not in item.name.casefold()
    ]
    if not candidates:
        candidates = [item for item in trial.interventions if "placebo" not in item.name.casefold()]
    selected = candidates[0] if candidates else None
    if len(candidates) > 1:
        flags.append(_missing("asset-multiple-candidates", "asset_identity", "asset_name", "Multiple non-placebo interventions need review.", "medium"))
    if selected is None:
        flags.append(_missing("asset-missing-name", "asset_identity", "asset_name", "No non-placebo asset candidate was found.", "high"))

    rxnorm_match = None
    if selected is not None:
        try:
            rxnorm_match, rx_source = (rxnorm_client or RxNormClient()).normalize(selected.name)
            sources.append(rx_source)
            if rxnorm_match is None:
                flags.append(_missing("asset-no-rxnorm", "asset_identity", "rxnorm_match", "RxNorm returned no match.", "medium"))
        except RxNormError as exc:
            flags.append(_missing("asset-rxnorm-error", "asset_identity", "rxnorm_match", str(exc), "medium"))

    modality, modality_rule = _infer_modality(selected)
    if overrides.get("modality"):
        modality = str(overrides["modality"])
        modality_rule = "human_override"
    indication, therapeutic_area, indication_rule = _infer_indication(trial)
    if overrides.get("indication"):
        indication = str(overrides["indication"])
        indication_rule = "human_override"
    if overrides.get("therapeutic_area"):
        therapeutic_area = str(overrides["therapeutic_area"])
        indication_rule = "human_override"
    sponsor = trial.lead_sponsor.name if trial.lead_sponsor else None
    sponsor_rule = "lead_sponsor_fallback" if sponsor else None
    aliases_config = load_rule_config("sponsor_aliases.yaml").get("aliases", {})
    if isinstance(aliases_config, dict) and sponsor in aliases_config:
        sponsor = str(aliases_config[sponsor])
        sponsor_rule = "sponsor_alias_exact"
    if overrides.get("sponsor"):
        sponsor = str(overrides["sponsor"])
        sponsor_rule = "human_override"
    if sponsor is None:
        flags.append(_missing("asset-missing-sponsor", "asset_identity", "sponsor", "ClinicalTrials.gov did not list a lead sponsor.", "medium"))
    if indication is None:
        flags.append(_missing("asset-missing-indication", "asset_identity", "normalized_indication", "No deterministic indication rule matched.", "medium"))
    if modality == "unknown":
        flags.append(_missing("asset-unknown-modality", "asset_identity", "modality", "No deterministic modality rule matched.", "medium"))

    aliases = tuple(dict.fromkeys([*(selected.other_names if selected else ()), *(rxnorm_match.aliases if rxnorm_match else ())]))
    confidence = 0.85 - min(len(flags), 4) * 0.15
    return (
        AssetIdentityOutput(
            nct_id=trial.nct_id,
            asset_name=selected.name if selected else None,
            raw_intervention_names=tuple(item.name for item in trial.interventions),
            intervention_type=selected.type if selected else None,
            aliases=aliases,
            rxnorm_match=rxnorm_match,
            sponsor=sponsor,
            normalized_indication=indication,
            therapeutic_area=therapeutic_area,
            modality=modality,
            rule_ids=tuple(item for item in (modality_rule, indication_rule, sponsor_rule) if item),
            source_ids=tuple(source.source_id for source in sources),
            missing_data_flags=tuple(flags),
            confidence=max(0.1, confidence),
        ),
        tuple(sources),
    )


def search_patent_exclusivity(
    asset: AssetIdentityOutput,
    *,
    loe_year_override: int | None = None,
    client: httpx.Client | None = None,
) -> tuple[PatentExclusivityOutput, tuple[SourceMetadata, ...]]:
    """Search Lens when configured; otherwise require human LOE review."""

    token = os.getenv("LENS_API_TOKEN")
    source = SourceMetadata(
        source_id=f"lens:{_slug(asset.asset_name or asset.nct_id)}",
        title=f"Lens patent search for {asset.asset_name or asset.nct_id}",
        url=LENS_PATENT_URL,
        provenance="Lens Patent Search API",
        source_type="patent_search",
        version="v1",
    )
    override_source = SourceMetadata(
        source_id=f"human_override:loe:{_slug(asset.nct_id)}",
        title=f"Reviewed LOE override for {asset.nct_id}",
        provenance="CLI supplied reviewed LOE year",
        source_type="human_override",
        version="local",
    )
    terms = tuple(item for item in (asset.asset_name, asset.sponsor, *asset.aliases[:3]) if item)
    flags: list[MissingDataFlag] = []
    candidates: list[PatentCandidate] = []
    if not token:
        flags.append(_missing("patent-lens-token-missing", "patent_exclusivity", "loe_year", "LENS_API_TOKEN is missing; Lens retrieval skipped.", "high"))
    elif not terms:
        flags.append(_missing("patent-search-terms-missing", "patent_exclusivity", "searched_terms", "No asset/sponsor terms were available for patent search.", "high"))
    else:
        try:
            response = (client or httpx.Client(timeout=20.0)).post(
                LENS_PATENT_URL,
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                json={
                    "query": {"query_string": {"query": " OR ".join(f'"{term}"' for term in terms)}},
                    "size": 5,
                    "from": 0,
                    "include": ["lens_id", "biblio.invention_title", "jurisdiction", "date_published", "legal_status"],
                },
                timeout=20.0,
            )
            if response.status_code not in {200, 204, 404}:
                flags.append(_missing("patent-lens-http-error", "patent_exclusivity", "candidates", f"Lens returned HTTP {response.status_code}.", "high"))
            elif response.status_code == 200:
                payload = response.json()
                for raw in _lens_records(payload):
                    candidate_id = str(raw.get("lens_id") or raw.get("doc_number") or "")
                    if not candidate_id:
                        continue
                    candidates.append(
                        PatentCandidate(
                            candidate_id=candidate_id,
                            title=_first_title(raw),
                            jurisdiction=_first_text(raw.get("jurisdiction")),
                            publication_date=_first_text(raw.get("date_published")),
                            legal_status=_first_text(raw.get("legal_status")),
                            source_id=source.source_id,
                        )
                    )
            if not candidates:
                flags.append(_missing("patent-no-candidates", "patent_exclusivity", "candidates", "No Lens patent candidates were retrieved.", "medium"))
        except Exception as exc:
            flags.append(_missing("patent-lens-error", "patent_exclusivity", "candidates", f"Lens request failed: {exc.__class__.__name__}.", "high"))

    if loe_year_override is None:
        flags.append(_missing("patent-loe-review-required", "patent_exclusivity", "estimated_loe_year", "No reviewed LOE year was supplied.", "high"))
    output_sources = []
    if token:
        output_sources.append(source)
    if loe_year_override is not None:
        output_sources.append(override_source)
    return (
        PatentExclusivityOutput(
            asset_name=asset.asset_name,
            searched_terms=terms,
            candidates=tuple(candidates),
            estimated_loe_year=loe_year_override,
            source_ids=tuple(item.source_id for item in output_sources),
            missing_data_flags=tuple(flags),
            confidence=0.75 if candidates and loe_year_override else 0.2,
        ),
        tuple(output_sources),
    )


def lookup_pos(
    trial: ClinicalTrialRecord,
    asset: AssetIdentityOutput,
    *,
    workbook_path: str | None = None,
) -> tuple[PoSOutput, SourceMetadata]:
    """Lookup source-only PoS from the local workbook."""

    path = Path(workbook_path or os.getenv("PHARMA_OS_POS_WORKBOOK_PATH") or DEFAULT_POS_WORKBOOK)
    source = SourceMetadata(
        source_id=f"pos_workbook:{_slug(path.name)}",
        title="Source-based probability of success workbook",
        url=None,
        provenance="Local Source_Based_PoS_Workbook.xlsx AllBenchmarks lookup",
        source_type="pos_workbook",
        version=path.name,
    )
    flags: list[MissingDataFlag] = []
    if not path.exists():
        return (
            PoSOutput(workbook_path=str(path), source_ids=(), missing_data_flags=(_missing("pos-workbook-missing", "pos", "workbook_path", f"Workbook not found: {path}", "high"),), confidence=0.0),
            source,
        )
    phase = _phase_for_workbook(trial.phases)
    disease_area = _disease_area_for_workbook(asset.therapeutic_area, trial.conditions)
    if not phase:
        flags.append(_missing("pos-phase-missing", "pos", "current_phase", "Trial phase could not be mapped to workbook labels.", "high"))
    if not disease_area:
        flags.append(_missing("pos-disease-area-missing", "pos", "disease_area", "Disease area could not be mapped to workbook labels.", "high"))
    value = None
    row_out: dict[str, str | int | float | bool | None] = {}
    lookup_key = f"Disease Area|{disease_area}|{phase}" if disease_area and phase else None
    if lookup_key:
        try:
            wb = load_workbook(path, data_only=False, read_only=False)
            ws = wb["AllBenchmarks"]
            headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_map = dict(zip(headers, row))
                if _norm(str(row_map.get("Key") or "")) == _norm(lookup_key):
                    value = _float(row_map.get("Phase LOA"))
                    row_out = {str(k): _json_scalar(v) for k, v in row_map.items() if k is not None}
                    break
            if value is None:
                flags.append(_missing("pos-row-missing", "pos", "probability_of_success", f"No workbook row found for {lookup_key}.", "high"))
        except Exception as exc:
            flags.append(_missing("pos-workbook-error", "pos", "probability_of_success", f"Workbook lookup failed: {exc.__class__.__name__}.", "high"))
    return (
        PoSOutput(
            probability_of_success=value,
            current_phase=phase,
            disease_area=disease_area,
            workbook_path=str(path),
            lookup_key=lookup_key,
            benchmark_row=row_out,
            source_ids=(source.source_id,) if value is not None else (),
            missing_data_flags=tuple(flags),
            confidence=0.9 if value is not None else 0.0,
        ),
        source,
    )


def lookup_pricing(
    asset: AssetIdentityOutput,
    *,
    wac_data_path: str | None = None,
    client: httpx.Client | None = None,
) -> tuple[PricingOutput, tuple[SourceMetadata, ...]]:
    """Match local WAC rows and openFDA dosing/label data."""

    wac_config = load_config("wac_sources.yaml", section="due_diligence")
    wac_config_source = config_source("wac_sources.yaml", section="due_diligence")
    configured_path = _configured_wac_path(wac_config)
    path = Path(wac_data_path or os.getenv("PHARMA_OS_WAC_DATA_PATH") or configured_path or DEFAULT_WAC_DATA)
    wac_source = SourceMetadata(
        source_id=f"wac:{_slug(path.name)}",
        title="Local WAC workbook",
        provenance=f"Local WAC workbook combined_wac_increases lookup approved by {config_provenance('wac_sources.yaml', 'approved_sources[0]', section='due_diligence')}",
        source_type="wac_workbook",
        version=path.name,
    )
    label_source = SourceMetadata(
        source_id=f"openfda_label:{_slug(asset.asset_name or asset.nct_id)}",
        title=f"openFDA label search for {asset.asset_name or asset.nct_id}",
        url=OPENFDA_LABEL_URL,
        provenance="openFDA drug label API",
        source_type="drug_label",
        version="openFDA",
    )
    flags: list[MissingDataFlag] = []
    wac_value = None
    matched_product = None
    if not path.exists():
        flags.append(_missing("pricing-wac-file-missing", "pricing", "wac_value", f"WAC workbook not found: {path}", "high"))
    elif not asset.asset_name:
        flags.append(_missing("pricing-asset-missing", "pricing", "matched_product", "No asset name available for WAC lookup.", "high"))
    else:
        try:
            terms = _pricing_terms(asset)
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb["combined_wac_increases"]
            headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_map = dict(zip(headers, row))
                description = str(row_map.get("Drug Product Description") or "")
                if any(_norm(term) and _norm(term) in _norm(description) for term in terms):
                    wac_value = _float(row_map.get("WAC After Increase"))
                    matched_product = description
                    break
            if wac_value is None:
                flags.append(_missing("pricing-no-wac-match", "pricing", "wac_value", f"No WAC row matched {asset.asset_name}.", "high"))
        except Exception as exc:
            flags.append(_missing("pricing-wac-error", "pricing", "wac_value", f"WAC lookup failed: {exc.__class__.__name__}.", "high"))

    dosing_summary = None
    terms = _pricing_terms(asset)
    if terms:
        try:
            label_payload = None
            for term in terms:
                response = (client or httpx.Client(timeout=20.0)).get(
                    OPENFDA_LABEL_URL,
                    params={"search": f'openfda.brand_name:"{term}"', "limit": "1"},
                    timeout=20.0,
                )
                if response.status_code == 200:
                    label_payload = response.json()
                    break
                response = (client or httpx.Client(timeout=20.0)).get(
                    OPENFDA_LABEL_URL,
                    params={"search": f'openfda.generic_name:"{term}"', "limit": "1"},
                    timeout=20.0,
                )
                if response.status_code == 200:
                    label_payload = response.json()
                    break
            if label_payload is not None:
                result = (label_payload.get("results") or [{}])[0]
                dosing_summary = _first_text(result.get("dosage_and_administration"))
            else:
                flags.append(_missing("pricing-no-openfda-label", "pricing", "dosing_summary", "openFDA returned no label match.", "medium"))
        except Exception as exc:
            flags.append(_missing("pricing-openfda-error", "pricing", "dosing_summary", f"openFDA lookup failed: {exc.__class__.__name__}.", "medium"))
    annual_wac = wac_value
    if wac_value is not None and dosing_summary is None:
        flags.append(_missing("pricing-dosing-review-required", "pricing", "annual_wac", "WAC was found but annualization lacks sourced dosing.", "high"))
    sources = tuple(source for source in (wac_config_source, wac_source if wac_value is not None else None, label_source if dosing_summary else None) if source)
    return (
        PricingOutput(
            annual_wac=annual_wac if dosing_summary else None,
            wac_value=wac_value,
            wac_unit_basis="package" if wac_value is not None else None,
            matched_product=matched_product,
            dosing_summary=dosing_summary,
            source_ids=tuple(source.source_id for source in sources),
            missing_data_flags=tuple(flags),
            confidence=0.8 if wac_value is not None and dosing_summary else 0.2,
        ),
        sources,
    )


def build_commercial_model(
    *,
    annual_patients: float | None,
    peak_penetration: float | None,
    gross_to_net: float | None,
    pricing: PricingOutput,
) -> CommercialModelOutput:
    """Run a compact deterministic commercial model when reviewed inputs exist."""

    flags: list[MissingDataFlag] = []
    archetype_name = "chronic_specialty_prevalence"
    config = load_config("default_archetypes.yaml", section="due_diligence")
    config_id = config_source_id("default_archetypes.yaml", section="due_diligence")
    archetype = (config.get("archetypes") or {}).get(archetype_name, {})
    selected_peak_penetration = _select_assumption_value(
        "commercial-peak-penetration",
        "peak_penetration",
        peak_penetration,
        "fraction",
        config_value=_triplet_base(archetype.get("peak_penetration")),
        config_filename="default_archetypes.yaml",
        config_field_path=f"archetypes.{archetype_name}.peak_penetration.base",
    )
    selected_gross_to_net = _select_assumption_value(
        "commercial-gross-to-net",
        "gross_to_net",
        gross_to_net,
        "fraction",
        config_value=_triplet_base(archetype.get("gross_to_net")),
        config_filename="default_archetypes.yaml",
        config_field_path=f"archetypes.{archetype_name}.gross_to_net.base",
    )
    launch_ramp = [float(value) for value in archetype.get("launch_ramp") or [] if _float(value) is not None]
    assumptions = [
        _assumption("commercial-annual-patients", "annual_patients", annual_patients, "patients", "cli.due_diligence", assumption_type="user_reviewed"),
        selected_peak_penetration,
        selected_gross_to_net,
    ]
    if launch_ramp:
        assumptions.append(
            _assumption(
                "commercial-launch-ramp",
                "launch_ramp",
                launch_ramp,
                "fraction_by_year",
                config_provenance("default_archetypes.yaml", f"archetypes.{archetype_name}.launch_ramp", section="due_diligence"),
                assumption_type="config_default",
                source_ids=(config_id,),
            )
        )
    else:
        flags.append(_missing("commercial-launch-ramp-missing", "commercial_model", "launch_ramp", "No launch ramp was available from default_archetypes.yaml.", "high"))
    for assumption in assumptions:
        if assumption.value is None:
            flags.append(_missing(f"{assumption.assumption_id}-missing", "commercial_model", assumption.name, "No source-backed, user-reviewed, or config fallback value is available.", "high"))
    if pricing.annual_wac is None:
        flags.append(_missing("commercial-annual-wac-missing", "commercial_model", "annual_wac", "Annual WAC must come from pricing evidence.", "high"))
    calculable = not flags
    rows: list[RevenueForecastYear] = []
    net_price = None
    peak_sales = None
    if calculable:
        net_price = float(pricing.annual_wac) * (1 - float(selected_gross_to_net.value))
        for year, ramp in enumerate(launch_ramp, start=1):
            treated = float(annual_patients) * float(selected_peak_penetration.value) * ramp
            rows.append(RevenueForecastYear(year=year, treated_patients=round(treated, 2), net_price=round(net_price, 2), net_revenue=round(treated * net_price, 2)))
        peak_sales = rows[-1].net_revenue
    return CommercialModelOutput(
        calculable=calculable,
        annual_patients=annual_patients,
        peak_penetration=float(selected_peak_penetration.value) if selected_peak_penetration.value is not None else None,
        gross_to_net=float(selected_gross_to_net.value) if selected_gross_to_net.value is not None else None,
        net_price=net_price,
        peak_net_sales=peak_sales,
        revenue_forecast=tuple(rows),
        assumptions=tuple(assumptions),
        source_ids=tuple(dict.fromkeys([*pricing.source_ids, *([config_id] if any(item.source_ids and config_id in item.source_ids for item in assumptions) else [])])),
        missing_data_flags=tuple(flags),
        confidence=0.75 if calculable else 0.15,
    )


def build_rnpv(
    *,
    commercial: CommercialModelOutput,
    pos: PoSOutput,
    patent: PatentExclusivityOutput,
    launch_year: int | None,
    loe_year: int | None,
    discount_rate: float | None,
    operating_margin: float | None,
    development_cost: float | None,
    phase: str | None = None,
) -> RNPVOutput:
    """Calculate rNPV when all upstream sourced/reviewed inputs exist."""

    flags: list[MissingDataFlag] = []
    config = load_config("rnpv_assumptions_config.yaml", section="due_diligence")
    config_id = config_source_id("rnpv_assumptions_config.yaml", section="due_diligence")
    selected_launch_year = _select_assumption_value(
        "rnpv-launch-year",
        "launch_year",
        launch_year,
        "year",
        config_value=_launch_year_from_config(config, phase),
        config_filename="rnpv_assumptions_config.yaml",
        config_field_path=f"launch_timing.default_years_to_launch_by_phase.{phase or 'default'}",
    )
    selected_discount_rate = _select_assumption_value(
        "rnpv-discount-rate",
        "discount_rate",
        discount_rate,
        "fraction",
        config_value=_float(config.get("discount_rate")),
        config_filename="rnpv_assumptions_config.yaml",
        config_field_path="discount_rate",
    )
    selected_operating_margin = _select_assumption_value(
        "rnpv-operating-margin",
        "operating_margin",
        operating_margin,
        "fraction",
        config_value=_float(config.get("operating_margin")),
        config_filename="rnpv_assumptions_config.yaml",
        config_field_path="operating_margin",
    )
    selected_development_cost = _select_assumption_value(
        "rnpv-development-cost",
        "development_cost",
        development_cost,
        "USD",
        config_value=_development_cost_from_config(config, phase),
        config_filename="rnpv_assumptions_config.yaml",
        config_field_path=f"development_costs.by_phase.{phase or 'default'}.total_cost",
    )
    selected_tax_rate = _select_assumption_value(
        "rnpv-tax-rate",
        "tax_rate",
        None,
        "fraction",
        config_value=_float(config.get("tax_rate")),
        config_filename="rnpv_assumptions_config.yaml",
        config_field_path="tax_rate",
    )
    selected_valuation_year = _select_assumption_value(
        "rnpv-valuation-year",
        "valuation_year",
        None,
        "year",
        config_value=_float(config.get("valuation_year")),
        config_filename="rnpv_assumptions_config.yaml",
        config_field_path="valuation_year",
    )
    selected_loe_year = loe_year or patent.estimated_loe_year
    assumptions = [
        selected_launch_year,
        _assumption(
            "rnpv-loe-year",
            "loe_year",
            selected_loe_year,
            "year",
            "cli.due_diligence" if loe_year is not None else "Lens/regulatory source",
            assumption_type="user_reviewed" if loe_year is not None else "source_derived",
            source_ids=() if loe_year is not None else patent.source_ids,
        ),
        selected_discount_rate,
        selected_operating_margin,
        selected_development_cost,
        selected_tax_rate,
        selected_valuation_year,
    ]
    for assumption in assumptions:
        if assumption.value is None:
            flags.append(_missing(f"{assumption.assumption_id}-missing", "rnpv", assumption.name, "Reviewed rNPV assumption is required.", "high"))
    if not commercial.calculable:
        flags.append(_missing("rnpv-commercial-not-calculable", "rnpv", "commercial_model", "Commercial model is not calculable.", "high"))
    if pos.probability_of_success is None:
        flags.append(_missing("rnpv-pos-missing", "rnpv", "probability_of_success", "PoS must come from workbook.", "high"))
    if selected_loe_year is None:
        flags.append(_missing("rnpv-loe-missing", "rnpv", "loe_year", "LOE must cite Lens/regulatory source or human review.", "high"))
    if selected_launch_year.value and selected_loe_year and selected_loe_year < int(float(selected_launch_year.value)):
        flags.append(_missing("rnpv-loe-before-launch", "rnpv", "loe_year", "LOE year is before launch year.", "critical"))
    calculable = not flags
    value = None
    if calculable:
        value = -float(selected_development_cost.value)
        for row in commercial.revenue_forecast:
            calendar_year = int(float(selected_launch_year.value)) + row.year - 1
            if calendar_year > int(selected_loe_year):
                continue
            years = max(0, calendar_year - int(float(selected_valuation_year.value)))
            cash_flow = row.net_revenue * float(selected_operating_margin.value) * (1 - float(selected_tax_rate.value)) * float(pos.probability_of_success)
            value += cash_flow / ((1 + float(selected_discount_rate.value)) ** years)
        value = round(value, 2)
    return RNPVOutput(
        calculable=calculable,
        rnpv=value,
        probability_of_success=pos.probability_of_success,
        loe_year=selected_loe_year,
        launch_year=int(float(selected_launch_year.value)) if selected_launch_year.value is not None else None,
        discount_rate=float(selected_discount_rate.value) if selected_discount_rate.value is not None else None,
        operating_margin=float(selected_operating_margin.value) if selected_operating_margin.value is not None else None,
        development_cost=float(selected_development_cost.value) if selected_development_cost.value is not None else None,
        assumptions=tuple(assumptions),
        source_ids=tuple(dict.fromkeys([*commercial.source_ids, *pos.source_ids, *patent.source_ids, *([config_id] if any(config_id in item.source_ids for item in assumptions) else [])])),
        missing_data_flags=tuple(flags),
        confidence=0.75 if calculable else 0.1,
    )


def _infer_modality(selected: Any) -> tuple[str, str | None]:
    text = " ".join([getattr(selected, "name", "") or "", getattr(selected, "description", "") or "", *getattr(selected, "other_names", ())]).casefold()
    config = load_rule_config("modality_rules.yaml")
    for rule in config.get("rules", []):
        keywords = [str(keyword).casefold() for keyword in rule.get("keywords", [])]
        if any(keyword in text for keyword in keywords):
            return str(rule.get("modality")), str(rule.get("id") or "modality_rule")
    return str(config.get("default", "unknown")), None


def _infer_indication(trial: ClinicalTrialRecord) -> tuple[str | None, str | None, str | None]:
    text = " | ".join([*trial.conditions, trial.brief_title or "", trial.official_title or ""]).casefold()
    for rule in load_rule_config("indication_rules.yaml").get("rules", []):
        terms = [str(term).casefold() for term in rule.get("terms", [])]
        if any(term in text for term in terms):
            return (
                str(rule.get("normalized_indication")),
                str(rule.get("therapeutic_area")),
                str(rule.get("id") or "indication_rule"),
            )
    if len(trial.conditions) == 1:
        return trial.conditions[0], None, None
    return None, None, None


def _pricing_terms(asset: AssetIdentityOutput) -> tuple[str, ...]:
    terms = [asset.asset_name, *asset.aliases]
    if asset.rxnorm_match:
        terms.extend([asset.rxnorm_match.matched_name, *asset.rxnorm_match.aliases])
    return tuple(dict.fromkeys(term for term in terms if term))


def _configured_wac_path(config: dict[str, Any]) -> str | None:
    sources = config.get("approved_sources")
    if not isinstance(sources, list):
        return None
    for source in sources:
        if isinstance(source, dict) and source.get("local_path"):
            return str(source["local_path"])
    return None


def _disease_area_for_workbook(therapeutic_area: str | None, conditions: tuple[str, ...]) -> str | None:
    text = " ".join([therapeutic_area or "", *conditions]).casefold()
    if "oncology" in text or "cancer" in text or "tumor" in text or "glioblastoma" in text:
        return "Oncology"
    if "neurology" in text or "alzheimer" in text or "parkinson" in text:
        return "Neurology"
    if "immunology" in text:
        return "Autoimmune"
    return None


def _phase_for_workbook(phases: tuple[str, ...]) -> str | None:
    text = " ".join(phases).upper()
    if "PHASE3" in text or "PHASE 3" in text:
        return "Phase III"
    if "PHASE2" in text or "PHASE 2" in text:
        return "Phase II"
    if "PHASE1" in text or "PHASE 1" in text or "EARLY_PHASE1" in text:
        return "Phase I"
    return None


def _missing(flag_id: str, section: str, field: str, reason: str, severity: str) -> MissingDataFlag:
    return MissingDataFlag(flag_id=flag_id, section=section, field=field, reason=reason, severity=severity)  # type: ignore[arg-type]


def _select_assumption_value(
    assumption_id: str,
    name: str,
    user_value: Any,
    unit: str,
    *,
    config_value: Any,
    config_filename: str,
    config_field_path: str,
) -> AssumptionRecord:
    if user_value is not None:
        return _assumption(
            assumption_id,
            name,
            user_value,
            unit,
            "cli.due_diligence",
            assumption_type="user_reviewed",
        )
    if config_value is not None:
        return _assumption(
            assumption_id,
            name,
            config_value,
            unit,
            config_provenance(config_filename, config_field_path, section="due_diligence"),
            assumption_type="fallback_assumption",
            source_ids=(config_source_id(config_filename, section="due_diligence"),),
        )
    return _assumption(
        assumption_id,
        name,
        None,
        unit,
        f"missing:{config_filename}:{config_field_path}",
        assumption_type="missing",
        requires_human_review=True,
    )


def _assumption(
    assumption_id: str,
    name: str,
    value: Any,
    unit: str,
    provenance: str,
    *,
    assumption_type: str,
    source_ids: tuple[str, ...] = (),
    requires_human_review: bool | None = None,
) -> AssumptionRecord:
    return AssumptionRecord(
        assumption_id=assumption_id,
        name=name,
        value=value,
        unit=unit,
        assumption_type=assumption_type,  # type: ignore[arg-type]
        source_ids=source_ids,
        provenance=provenance,
        requires_human_review=value is None if requires_human_review is None else requires_human_review,
    )


def _triplet_base(value: Any) -> float | None:
    return _float(value.get("base")) if isinstance(value, dict) else None


def _launch_year_from_config(config: dict[str, Any], phase: str | None) -> int | None:
    valuation_year = _float(config.get("valuation_year"))
    timing = ((config.get("launch_timing") or {}).get("default_years_to_launch_by_phase") or {})
    years = _float(timing.get(phase or "") or timing.get("default"))
    if valuation_year is None or years is None:
        return None
    return int(valuation_year + years)


def _development_cost_from_config(config: dict[str, Any], phase: str | None) -> float | None:
    by_phase = ((config.get("development_costs") or {}).get("by_phase") or {})
    selected = by_phase.get(phase or "") or by_phase.get("default")
    if not isinstance(selected, dict):
        return None
    return _float(selected.get("total_cost"))


def _lens_records(payload: Any) -> list[dict[str, Any]]:
    if not isinstance(payload, dict):
        return []
    data = payload.get("data") or payload.get("results") or []
    if isinstance(data, dict):
        data = data.get("data") or data.get("results") or []
    return [item for item in data if isinstance(item, dict)]


def _first_title(raw: dict[str, Any]) -> str | None:
    biblio = raw.get("biblio") if isinstance(raw.get("biblio"), dict) else {}
    title = raw.get("title") or biblio.get("invention_title")
    return _first_text(title)


def _first_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, list):
        return _first_text(value[0]) if value else None
    if isinstance(value, dict):
        for item in value.values():
            text = _first_text(item)
            if text:
                return text
        return None
    text = str(value).strip()
    return text or None


def _float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number / 100 if number > 1 and number <= 100 and "%" in str(value) else number


def _json_scalar(value: Any) -> str | int | float | bool | None:
    if value is None or isinstance(value, (str, int, float, bool)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return value
    return str(value)


def _norm(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.casefold()).strip("-") or "unknown"
